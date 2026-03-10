import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_quotation_excel(project_name: str, items: List[Dict[str, Any]]) -> bytes:
    """Generate a professionally formatted Excel quotation file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="1a1a2e")
    currency_format = 'R$ #,##0.00'
    pct_format = '0.00"%"'
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Orçamento — {project_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"] = "Gerado por Preço Inteligente"
    ws["A2"].font = Font(name="Calibri", size=10, color="888888")
    ws.row_dimensions[2].height = 20

    # Empty row
    ws.row_dimensions[3].height = 10

    # Headers
    headers = ["Produto", "Qtd", "Custo Unit.", "Margem", "Preço Sugerido", "Total", "Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 30

    # Data rows
    row_num = 5
    total_cost = 0
    total_suggested = 0
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    for idx, item in enumerate(items):
        is_alt = idx % 2 == 1
        fill = alt_fill if is_alt else PatternFill(fill_type=None)

        qty = item.get("quantity", 1)
        cost = item.get("cost", 0)
        margin = item.get("margin", 0)
        suggested = item.get("suggested_price", 0)
        line_total = suggested * qty
        total_cost += cost * qty
        total_suggested += line_total

        row_data = [
            item.get("product_name", ""),
            qty,
            cost,
            margin,
            suggested,
            line_total,
            item.get("product_url", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center
            if is_alt:
                cell.fill = fill

        # Format currency columns
        ws.cell(row=row_num, column=3).number_format = currency_format
        ws.cell(row=row_num, column=5).number_format = currency_format
        ws.cell(row=row_num, column=6).number_format = currency_format
        ws.cell(row=row_num, column=4).number_format = pct_format

        row_num += 1

    # Totals row
    row_num += 1
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=11)

    ws.cell(row=row_num, column=1, value="TOTAL").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=6, value=total_suggested).font = total_font
    ws.cell(row=row_num, column=6).number_format = currency_format
    ws.cell(row=row_num, column=6).fill = total_fill
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).fill = total_fill
        ws.cell(row=row_num, column=col).border = thin_border

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
